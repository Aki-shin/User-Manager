"""Parse XLSX employee file and produce normalized user records."""

import re
from openpyxl import load_workbook

# Expected column headers (case-insensitive matching)
EXPECTED_HEADERS = [
    'должность', 'подразделение', 'код', 'мобильный телефон',
    'рабочий телефон', 'email', 'имя', 'отчество', 'фамилия',
]

# Transliteration table (Russian → Latin)
TRANSLIT_MAP = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
    'ё': 'e', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
    'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
    'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts',
    'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ъ': '', 'ы': 'y', 'ь': '',
    'э': 'e', 'ю': 'yu', 'я': 'ya',
}


def transliterate(text):
    """Transliterate Russian text to Latin."""
    result = []
    for ch in text.lower():
        result.append(TRANSLIT_MAP.get(ch, ch))
    return ''.join(result)


def digits_only(value):
    """Strip everything except digits from a string."""
    if not value:
        return ''
    return re.sub(r'\D', '', str(value))


def generate_uid(surname, firstname, patronymic, existing_uids):
    """Generate uid from transliterated first letters of ФИО.

    Format: first letter of surname + first letter of firstname + first letter of patronymic.
    Append digit if collision.
    """
    s = transliterate(surname)[0] if surname else ''
    f = transliterate(firstname)[0] if firstname else ''
    p = transliterate(patronymic)[0] if patronymic else ''
    base = (s + f + p).lower()
    if not base:
        base = 'user'

    candidate = base
    counter = 1
    while candidate in existing_uids:
        candidate = f'{base}{counter}'
        counter += 1
    return candidate


def parse_xlsx(file_path_or_stream):
    """Parse XLSX and return list of merged employee dicts keyed by 'код'.

    Returns list of dicts with keys:
        code, firstname, patronymic, surname, title, department,
        mobile, phone, email, cn, givenname, sn
    """
    wb = load_workbook(file_path_or_stream, read_only=True, data_only=True)
    ws = wb.active

    # Detect header row
    header_map = {}
    rows = ws.iter_rows()
    for row in rows:
        values = [str(c.value).strip().lower() if c.value else '' for c in row]
        matched = 0
        for i, v in enumerate(values):
            for eh in EXPECTED_HEADERS:
                if eh in v:
                    header_map[eh] = i
                    matched += 1
                    break
        if matched >= 5:
            break

    if len(header_map) < 5:
        raise ValueError('Не удалось распознать заголовки в XLSX файле. '
                         'Ожидаемые столбцы: ' + ', '.join(EXPECTED_HEADERS))

    def cell(row, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        v = row[idx].value
        return str(v).strip() if v is not None else ''

    # Collect rows grouped by code
    grouped = {}
    for row in rows:
        code = cell(row, 'код').strip()
        if not code:
            continue
        entry = {
            'title': cell(row, 'должность'),
            'department': cell(row, 'подразделение'),
            'mobile': cell(row, 'мобильный телефон'),
            'phone': cell(row, 'рабочий телефон'),
            'email': cell(row, 'email'),
            'firstname': cell(row, 'имя'),
            'patronymic': cell(row, 'отчество'),
            'surname': cell(row, 'фамилия'),
        }
        if code in grouped:
            grouped[code].append(entry)
        else:
            grouped[code] = [entry]

    wb.close()

    # Merge multiple rows per code
    result = []
    for code, entries in grouped.items():
        first = entries[0]
        titles = []
        departments = []
        for e in entries:
            if e['title'] and e['title'] not in titles:
                titles.append(e['title'])
            if e['department'] and e['department'] not in departments:
                departments.append(e['department'])

        firstname = first['firstname']
        patronymic = first['patronymic']
        surname = first['surname']

        givenname = f"{firstname} {patronymic}".strip() if patronymic else firstname
        sn = surname
        cn = f"{surname} {firstname} {patronymic}".strip()

        result.append({
            'code': code,
            'firstname': firstname,
            'patronymic': patronymic,
            'surname': surname,
            'givenname': givenname,
            'sn': sn,
            'cn': cn,
            'title': ' / '.join(titles),
            'department': ' / '.join(departments),
            'mobile': digits_only(first['mobile']),
            'phone': digits_only(first['phone']),
            'email': re.sub(r'\s+', '', first['email']),
        })

    return result
